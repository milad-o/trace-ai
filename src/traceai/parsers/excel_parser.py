"""Excel parser for workbooks containing data flows, formulas, and data lineage.

Parses Excel workbooks that define:
- Data transformation workflows
- Formula dependencies
- Sheet-to-sheet data flow
- Data mappings and lookups
"""

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from traceai.parsers.base import (
    BaseParser,
    Component,
    DataEntity,
    DataSource,
    Dependency,
    DocumentMetadata,
    DocumentType,
    Parameter,
    ParsedDocument,
)


class ExcelParser(BaseParser):
    """Parser for Excel workbooks."""

    def supported_extensions(self) -> list[str]:
        """Returns list of supported file extensions."""
        return [".xlsx", ".xlsm"]

    def document_type(self) -> DocumentType:
        """Returns the document type this parser handles."""
        return DocumentType.EXCEL_WORKBOOK

    def validate(self, file_path: Path) -> bool:
        """Validates if the file is a valid Excel workbook."""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            workbook.close()
            return True
        except Exception:
            return False

    def parse(self, file_path: Path) -> ParsedDocument:
        """Parses an Excel workbook.

        Extracts:
        - Sheets as components
        - Named ranges as parameters
        - Tables/ranges as data entities
        - Formula dependencies between sheets
        """
        workbook = openpyxl.load_workbook(file_path, data_only=False)

        # Extract metadata
        metadata = self._extract_metadata(file_path, workbook)

        # Parse components (sheets)
        components: list[Component] = []
        dependencies: list[Dependency] = []
        data_entities: list[DataEntity] = []
        parameters: list[Parameter] = []

        # Process each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_id = f"{metadata.document_id}_sheet_{sheet_name}"

            # Create component for sheet
            component = Component(
                name=sheet_name,
                component_id=sheet_id,
                component_type="excel_sheet",
                description=f"Excel sheet: {sheet_name}",
                properties={
                    "row_count": sheet.max_row,
                    "column_count": sheet.max_column,
                    "has_formulas": self._has_formulas(sheet),
                },
            )
            components.append(component)

            # Extract formula dependencies
            sheet_deps = self._extract_formula_dependencies(sheet, sheet_id, metadata.document_id)
            dependencies.extend(sheet_deps)

        # Extract named ranges as parameters
        if workbook.defined_names:
            for name in workbook.defined_names.definedName:
                param = Parameter(
                    name=name.name,
                    value=str(name.value),
                    description=f"Named range: {name.value}",
                )
                parameters.append(param)

        # Extract tables as data entities
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if hasattr(sheet, "tables") and sheet.tables:
                for table_name, table in sheet.tables.items():
                    entity = DataEntity(
                        name=table_name,
                        entity_type="excel_table",
                        description=f"Excel table in {sheet_name}",
                        properties={
                            "sheet": sheet_name,
                            "range": table.ref,
                        },
                    )
                    data_entities.append(entity)

        workbook.close()

        return ParsedDocument(
            metadata=metadata,
            components=components,
            data_sources=[],
            parameters=parameters,
            data_entities=data_entities,
            dependencies=dependencies,
        )

    def _extract_metadata(self, file_path: Path, workbook: Any) -> DocumentMetadata:
        """Extracts metadata from Excel workbook."""
        props = workbook.properties

        return DocumentMetadata(
            name=file_path.stem,
            document_id=f"excel_{file_path.stem}",
            document_type=self.document_type(),
            description=props.description if props else None,
            version=props.version if props else None,
            creator=props.creator if props else None,
            created_date=str(props.created) if props and props.created else None,
            modified_date=str(props.modified) if props and props.modified else None,
            file_path=file_path,
            custom_attributes={
                "sheet_count": len(workbook.sheetnames),
                "sheets": workbook.sheetnames,
            },
        )

    def _has_formulas(self, sheet: Worksheet) -> bool:
        """Checks if sheet contains any formulas."""
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    return True
        return False

    def _extract_formula_dependencies(
        self, sheet: Worksheet, sheet_id: str, doc_id: str
    ) -> list[Dependency]:
        """Extracts dependencies from formulas that reference other sheets."""
        dependencies: list[Dependency] = []
        referenced_sheets: set[str] = set()

        # Scan all cells for formulas
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    # Extract sheet references from formula
                    # Simple pattern: looks for 'SheetName!'
                    formula = cell.value
                    parts = formula.split("!")
                    for i, part in enumerate(parts[:-1]):  # Exclude last part
                        # Get sheet name (last token before !)
                        tokens = part.split("'")
                        if len(tokens) >= 2:
                            # Sheet name in quotes: 'Sheet Name'!
                            ref_sheet = tokens[-1]
                        else:
                            # Sheet name without quotes: SheetName!
                            tokens = part.split()
                            if tokens:
                                ref_sheet = tokens[-1]
                            else:
                                continue

                        if ref_sheet and ref_sheet != sheet.title:
                            referenced_sheets.add(ref_sheet)

        # Create dependencies for each referenced sheet
        for ref_sheet in referenced_sheets:
            ref_sheet_id = f"{doc_id}_sheet_{ref_sheet}"
            dependencies.append(
                Dependency(
                    from_id=ref_sheet_id,
                    to_id=sheet_id,
                    dependency_type="formula_reference",
                )
            )

        return dependencies

    def extract_data_entities(self, component: Component) -> list[DataEntity]:
        """Extracts data entities from Excel sheet component."""
        # Excel sheets themselves can be considered data entities
        return [
            DataEntity(
                name=component.name,
                entity_type="excel_sheet",
                description=component.description,
                properties=component.properties,
            )
        ]
