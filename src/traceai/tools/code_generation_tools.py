"""Code generation tools for exporting knowledge graph data.

Provides tools to generate:
- JSON exports of graph data
- CSV lineage reports
- Excel workbooks with analysis results
- Python code from COBOL/JCL
"""

import json
from pathlib import Path
from typing import Any

import networkx as nx
import pandas as pd
from langchain.tools import BaseTool
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field

from traceai.graph.queries import GraphQueries


class GenerateJSONInput(BaseModel):
    """Input for JSON generation tool."""

    output_path: str = Field(description="Path where JSON file should be saved")
    include_metadata: bool = Field(
        default=True, description="Whether to include node metadata in output"
    )
    filter_node_type: str | None = Field(
        default=None, description="Optional: Filter to specific node type (package, task, table)"
    )


class GenerateJSONTool(BaseTool):
    """Tool to generate JSON export of knowledge graph."""

    name: str = "generate_json"
    description: str = """Generates a JSON file containing knowledge graph data.

    Use this to export:
    - Complete graph structure
    - Node and edge information
    - Metadata and properties
    - Filtered subsets of the graph

    Args:
        output_path: Where to save the JSON file (e.g., "output/graph_export.json")
        include_metadata: Include full metadata (default: true)
        filter_node_type: Only include specific node types (optional)

    Returns:
        Success message with file path and statistics
    """
    args_schema: type[BaseModel] = GenerateJSONInput
    queries: GraphQueries

    def __init__(self, queries: GraphQueries | nx.DiGraph):
        # Accept both GraphQueries and raw DiGraph for flexibility
        if isinstance(queries, nx.DiGraph):
            queries = GraphQueries(queries)
        super().__init__(queries=queries)

    def _run(
        self, output_path: str, include_metadata: bool = True, filter_node_type: str | None = None
    ) -> str:
        """Generates JSON export of graph."""
        graph = self.queries.graph

        # Collect nodes
        nodes = []
        for node_id, node_data in graph.nodes(data=True):
            if filter_node_type and node_data.get("type") != filter_node_type:
                continue

            node_info: dict[str, Any] = {"id": node_id, "type": node_data.get("type")}

            if include_metadata:
                node_info["name"] = node_data.get("name")
                node_info["properties"] = {
                    k: v for k, v in node_data.items() if k not in ["type", "name"]
                }
            else:
                node_info["name"] = node_data.get("name")

            nodes.append(node_info)

        # Collect edges
        edges = []
        for source, target, edge_data in graph.edges(data=True):
            # Filter edges if node filter is active
            if filter_node_type:
                source_type = graph.nodes[source].get("type")
                target_type = graph.nodes[target].get("type")
                if source_type != filter_node_type and target_type != filter_node_type:
                    continue

            edge_info = {
                "source": source,
                "target": target,
                "type": edge_data.get("edge_type", "unknown"),
            }

            if include_metadata and "properties" in edge_data:
                edge_info["properties"] = edge_data["properties"]

            edges.append(edge_info)

        # Create export structure (tests expect nodes/edges at top level + metadata)
        export_data = {
            "nodes": nodes,
            "edges": edges,
            "metadata": {
                "total_nodes": len(nodes),
                "total_edges": len(edges),
                "node_types": self._count_node_types(nodes),
            },
        }

        # Write to file
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(export_data, f, indent=2)

        return f"✅ JSON export successfully saved to {output_path}\n📊 {len(nodes)} nodes, {len(edges)} edges"

    def _count_node_types(self, nodes: list[dict]) -> dict[str, int]:
        """Counts nodes by type."""
        counts: dict[str, int] = {}
        for node in nodes:
            node_type = node.get("type", "unknown")
            counts[node_type] = counts.get(node_type, 0) + 1
        return counts


class GenerateCSVInput(BaseModel):
    """Input for CSV generation tool."""

    output_path: str = Field(description="Path where CSV file should be saved")
    export_type: str = Field(
        description="Type of export: 'lineage' (table lineage), 'nodes' (all nodes), or 'edges' (all relationships)"
    )


class GenerateCSVTool(BaseTool):
    """Tool to generate CSV exports of graph data."""

    name: str = "generate_csv"
    description: str = """Generates CSV file with knowledge graph data.

    Export types:
    - 'lineage': Data lineage (table -> table relationships)
    - 'nodes': All nodes with properties
    - 'edges': All edges/relationships

    Args:
        output_path: Where to save CSV (e.g., "output/lineage.csv")
        export_type: What to export ('lineage', 'nodes', or 'edges')

    Returns:
        Success message with file path and row count
    """
    args_schema: type[BaseModel] = GenerateCSVInput
    queries: GraphQueries

    def __init__(self, queries: GraphQueries | nx.DiGraph):
        # Accept both GraphQueries and raw DiGraph for flexibility
        if isinstance(queries, nx.DiGraph):
            queries = GraphQueries(queries)
        super().__init__(queries=queries)

    def _run(self, output_path: str, export_type: str = "lineage") -> str:
        """Generates CSV export."""
        if export_type == "lineage":
            df = self._generate_lineage_csv()
        elif export_type == "nodes":
            df = self._generate_nodes_csv()
        elif export_type == "edges":
            df = self._generate_edges_csv()
        else:
            return f"❌ Unknown export_type: {export_type}. Use 'lineage', 'nodes', or 'edges'"

        # Save to file
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(output_file, index=False)

        return f"✅ CSV export successfully saved to {output_path}\n📊 {len(df)} rows exported"

    def _generate_lineage_csv(self) -> pd.DataFrame:
        """Generates lineage mapping CSV."""
        graph = self.queries.graph
        rows = []

        for source, target, edge_data in graph.edges(data=True):
            source_data = graph.nodes[source]
            target_data = graph.nodes[target]

            # Include all data flow edges (tests expect from_id/to_id columns)
            edge_type = edge_data.get("edge_type", "unknown")
            if "READS" in str(edge_type).upper() or "WRITES" in str(edge_type).upper() or "TABLE" in str(source_data.get("node_type", "")).upper():
                rows.append(
                    {
                        "from_id": source,
                        "to_id": target,
                        "source": source_data.get("name", source),
                        "target": target_data.get("name", target),
                        "relationship_type": str(edge_type),
                        "transformation": edge_data.get("properties", {}).get(
                            "transformation", ""
                        ),
                    }
                )

        return pd.DataFrame(rows)

    def _generate_nodes_csv(self) -> pd.DataFrame:
        """Generates all nodes CSV."""
        graph = self.queries.graph
        rows = []

        for node_id, node_data in graph.nodes(data=True):
            rows.append(
                {
                    "node_id": node_id,  # Tests expect node_id column
                    "id": node_id,
                    "name": node_data.get("name", node_id),
                    "type": node_data.get("node_type", node_data.get("type", "unknown")),
                    "description": node_data.get("description", ""),
                }
            )

        return pd.DataFrame(rows)

    def _generate_edges_csv(self) -> pd.DataFrame:
        """Generates all edges CSV."""
        graph = self.queries.graph
        rows = []

        for source, target, edge_data in graph.edges(data=True):
            rows.append(
                {
                    "source": source,
                    "target": target,
                    "relationship_type": edge_data.get("edge_type", "unknown"),
                    "source_name": graph.nodes[source].get("name", source),
                    "target_name": graph.nodes[target].get("name", target),
                }
            )

        return pd.DataFrame(rows)


class GenerateExcelInput(BaseModel):
    """Input for Excel generation tool."""

    output_path: str = Field(description="Path where Excel file should be saved")
    include_sheets: list[str] = Field(
        default=["summary", "nodes", "lineage"],
        description="Sheets to include: 'summary', 'nodes', 'edges', 'lineage'",
    )


class GenerateExcelTool(BaseTool):
    """Tool to generate Excel workbook with graph analysis."""

    name: str = "generate_excel"
    description: str = """Generates Excel workbook with multiple sheets of graph data.

    Available sheets:
    - 'summary': Graph statistics and overview
    - 'nodes': All nodes with properties
    - 'edges': All relationships
    - 'lineage': Data lineage mappings

    Args:
        output_path: Where to save Excel file (e.g., "output/analysis.xlsx")
        include_sheets: Which sheets to include (default: all)

    Returns:
        Success message with file path and sheet count
    """
    args_schema: type[BaseModel] = GenerateExcelInput
    queries: GraphQueries

    def __init__(self, queries: GraphQueries | nx.DiGraph):
        # Accept both GraphQueries and raw DiGraph for flexibility
        if isinstance(queries, nx.DiGraph):
            queries = GraphQueries(queries)
        super().__init__(queries=queries)

    def _run(self, output_path: str, include_sheets: list[str] | None = None) -> str:
        """Generates Excel workbook."""
        if include_sheets is None:
            include_sheets = ["summary", "nodes", "lineage"]

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Add requested sheets
        if "summary" in include_sheets:
            self._add_summary_sheet(wb)
        if "nodes" in include_sheets:
            self._add_nodes_sheet(wb)
        if "edges" in include_sheets:
            self._add_edges_sheet(wb)
        if "lineage" in include_sheets:
            self._add_lineage_sheet(wb)

        # Save workbook
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)

        return (
            f"✅ Excel workbook successfully saved to {output_path}\n📊 {len(wb.sheetnames)} sheets created"
        )

    def _add_summary_sheet(self, wb: Workbook) -> None:
        """Adds summary sheet with graph statistics."""
        ws = wb.create_sheet("Summary")
        stats = self.queries.get_graph_stats()

        # Add header
        ws["A1"] = "Knowledge Graph Summary"
        ws["A1"].font = Font(size=14, bold=True)

        # Add statistics
        row = 3
        for key, value in stats.items():
            ws[f"A{row}"] = key
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Style
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _add_nodes_sheet(self, wb: Workbook) -> None:
        """Adds nodes sheet."""
        ws = wb.create_sheet("Nodes")
        headers = ["ID", "Name", "Type", "Description"]

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", fill_type="solid")

        # Add data
        graph = self.queries.graph
        row = 2
        for node_id, node_data in graph.nodes(data=True):
            ws.cell(row, 1, node_id)
            ws.cell(row, 2, node_data.get("name", ""))
            ws.cell(row, 3, node_data.get("type", ""))
            ws.cell(row, 4, node_data.get("description", ""))
            row += 1

        # Auto-size columns
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 20

    def _add_edges_sheet(self, wb: Workbook) -> None:
        """Adds edges sheet."""
        ws = wb.create_sheet("Edges")
        headers = ["Source", "Target", "Relationship", "Source Name", "Target Name"]

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", fill_type="solid")

        # Add data
        graph = self.queries.graph
        row = 2
        for source, target, edge_data in graph.edges(data=True):
            ws.cell(row, 1, source)
            ws.cell(row, 2, target)
            ws.cell(row, 3, edge_data.get("edge_type", ""))
            ws.cell(row, 4, graph.nodes[source].get("name", ""))
            ws.cell(row, 5, graph.nodes[target].get("name", ""))
            row += 1

        # Auto-size columns
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 20

    def _add_lineage_sheet(self, wb: Workbook) -> None:
        """Adds lineage sheet."""
        ws = wb.create_sheet("Lineage")
        headers = ["Source Table", "Target Table", "Relationship", "Transformation"]

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", fill_type="solid")

        # Add data
        graph = self.queries.graph
        row = 2
        for source, target, edge_data in graph.edges(data=True):
            source_data = graph.nodes[source]
            target_data = graph.nodes[target]

            if source_data.get("type") == "table" and target_data.get("type") == "table":
                ws.cell(row, 1, source_data.get("name", source))
                ws.cell(row, 2, target_data.get("name", target))
                ws.cell(row, 3, edge_data.get("edge_type", ""))
                ws.cell(row, 4, edge_data.get("properties", {}).get("transformation", ""))
                row += 1

        # Auto-size columns
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 25
