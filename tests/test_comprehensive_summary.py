"""Comprehensive summary test - validates all major components work."""

import asyncio
import tempfile
from pathlib import Path

import pytest

from traceai.agents import TraceAI
from traceai.graph.queries import GraphQueries
from traceai.parsers import parser_registry


class TestComprehensiveSummary:
    """High-level integration test covering all major functionality."""

    def test_all_parsers_registered(self):
        """Test all parsers are properly registered."""
        # Should find parsers for all supported formats
        assert parser_registry.get_parser_for_file(Path("test.dtsx")) is not None  # SSIS
        assert parser_registry.get_parser_for_file(Path("test.json")) is not None  # JSON
        assert parser_registry.get_parser_for_file(Path("test.csv")) is not None   # CSV
        assert parser_registry.get_parser_for_file(Path("test.xlsx")) is not None  # Excel
        assert parser_registry.get_parser_for_file(Path("test.cbl")) is not None   # COBOL
        assert parser_registry.get_parser_for_file(Path("test.jcl")) is not None   # JCL

    def test_traceai_end_to_end_sync(self):
        """Test TraceAI workflow using synchronous orchestration."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # 1. Create agent
            agent = TraceAI(persist_dir=temp_dir)

            # 2. Load documents
            ssis_dir = Path(__file__).parent.parent / "examples" / "inputs" / "ssis"
            if ssis_dir.exists():
                asyncio.run(agent.load_documents(ssis_dir))

                # 3. Verify graph was built
                assert agent.graph is not None
                assert agent.graph.number_of_nodes() > 0
                assert agent.graph.number_of_edges() > 0

                # 4. Verify documents were parsed
                assert len(agent.parsed_documents) > 0

                # 5. Test graph queries
                queries = GraphQueries(agent.graph)
                stats = queries.get_graph_stats()
                assert stats["total_nodes"] > 0
                assert stats["total_edges"] > 0

                # 6. Test lineage tracing
                lineage = queries.trace_data_lineage("Customer", direction="both")
                assert "upstream" in lineage
                assert "downstream" in lineage

                # 7. Test vector store
                results = agent.vector_store.similarity_search("customer", k=5)
                assert len(results) > 0

    @pytest.mark.asyncio
    async def test_traceai_end_to_end_async(self):
        """Test TraceAI workflow using native async operations."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # 1. Create async agent
            agent = TraceAI(persist_dir=temp_dir, max_concurrent_parsers=10)

            # 2. Load documents asynchronously
            ssis_dir = Path(__file__).parent.parent / "examples" / "inputs" / "ssis"
            if ssis_dir.exists():
                await agent.load_documents(ssis_dir)

                # 3. Verify graph was built
                assert agent.graph is not None
                assert agent.graph.number_of_nodes() > 0

                # 4. Verify documents were parsed
                assert len(agent.parsed_documents) > 0

                # 5. Test graph queries (sync operations on async agent)
                stats = agent.get_graph_stats()
                assert stats["total_nodes"] > 0

    def test_parser_properties(self):
        """Test all parsers have required properties."""
        from traceai.parsers.ssis_parser import SSISParser
        from traceai.parsers.json_parser import JSONParser
        from traceai.parsers.csv_parser import CSVParser
        from traceai.parsers.excel_parser import ExcelParser
        from traceai.parsers.cobol_parser import COBOLParser
        from traceai.parsers.jcl_parser import JCLParser

        parsers = [
            SSISParser(),
            JSONParser(),
            CSVParser(),
            ExcelParser(),
            COBOLParser(),
            JCLParser(),
        ]

        for parser in parsers:
            # Should have @property decorators
            assert hasattr(parser, "supported_extensions")
            assert hasattr(parser, "document_type")

            # Should return correct types
            exts = parser.supported_extensions
            assert isinstance(exts, list)
            assert len(exts) > 0
            assert all(isinstance(ext, str) for ext in exts)

            # Document type should be accessible
            doc_type = parser.document_type
            assert doc_type is not None

    def test_graph_tools_available(self):
        """Test graph tools can be created."""
        with tempfile.TemporaryDirectory() as temp_dir:
            agent = TraceAI(persist_dir=temp_dir)

            ssis_dir = Path(__file__).parent.parent / "examples" / "inputs" / "ssis"
            if ssis_dir.exists():
                asyncio.run(agent.load_documents(ssis_dir))

                from traceai.tools import create_graph_tools, create_graph_visualization_tool

                # Should be able to create tools
                graph_tools = create_graph_tools(agent.graph)
                assert len(graph_tools) > 0

                viz_tool = create_graph_visualization_tool(agent.graph)
                assert viz_tool is not None

    def test_code_generation_tools_available(self):
        """Test code generation tools can be instantiated."""
        with tempfile.TemporaryDirectory() as temp_dir:
            agent = TraceAI(persist_dir=temp_dir)

            ssis_dir = Path(__file__).parent.parent / "examples" / "inputs" / "ssis"
            if ssis_dir.exists():
                asyncio.run(agent.load_documents(ssis_dir))

                from traceai.tools.code_generation_tools import (
                    GenerateJSONTool,
                    GenerateCSVTool,
                    GenerateExcelTool,
                )
                from traceai.tools.python_generator import GeneratePythonTool

                # Should be able to create all tools
                json_tool = GenerateJSONTool(agent.graph)
                assert json_tool.name == "generate_json"

                csv_tool = GenerateCSVTool(agent.graph)
                assert csv_tool.name == "generate_csv"

                excel_tool = GenerateExcelTool(agent.graph)
                assert excel_tool.name == "generate_excel"

                python_tool = GeneratePythonTool()
                assert python_tool.name == "generate_python_from_cobol"

    def test_export_formats_work(self):
        """Test all export formats can be generated."""
        import json
        import pandas as pd
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            agent = TraceAI(persist_dir=temp_dir)

            ssis_dir = Path(__file__).parent.parent / "examples" / "inputs" / "ssis"
            if ssis_dir.exists():
                asyncio.run(agent.load_documents(ssis_dir))

                from traceai.tools.code_generation_tools import (
                    GenerateJSONTool,
                    GenerateCSVTool,
                    GenerateExcelTool,
                )

                output_dir = Path(temp_dir) / "output"
                output_dir.mkdir()

                # Test JSON export
                json_tool = GenerateJSONTool(agent.graph)
                json_path = output_dir / "export.json"
                json_tool._run(str(json_path))
                assert json_path.exists()
                with open(json_path) as f:
                    data = json.load(f)
                    assert "nodes" in data

                # Test CSV export
                csv_tool = GenerateCSVTool(agent.graph)
                csv_path = output_dir / "export.csv"
                csv_tool._run(str(csv_path))
                assert csv_path.exists()
                df = pd.read_csv(csv_path)
                assert len(df) > 0

                # Test Excel export
                excel_tool = GenerateExcelTool(agent.graph)
                excel_path = output_dir / "report.xlsx"
                excel_tool._run(str(excel_path))
                assert excel_path.exists()
                wb = load_workbook(excel_path)
                assert len(wb.sheetnames) > 0
