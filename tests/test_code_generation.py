"""Comprehensive tests for code generation tools."""

import json
import tempfile
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from traceai.agents import EnterpriseAgent
from traceai.tools.code_generation_tools import (
    GenerateJSONTool,
    GenerateCSVTool,
    GenerateExcelTool,
)
from traceai.tools.python_generator import GeneratePythonTool


@pytest.fixture
def agent_with_data():
    """Create an agent with loaded sample data."""
    with tempfile.TemporaryDirectory() as temp_dir:
        agent = EnterpriseAgent(persist_dir=temp_dir)

        # Load sample SSIS packages
        ssis_dir = Path(__file__).parent.parent / "examples" / "inputs" / "ssis"
        if ssis_dir.exists():
            agent.load_documents(ssis_dir)

        yield agent


@pytest.fixture
def temp_output_dir():
    """Create temporary output directory."""
    with tempfile.TemporaryDirectory() as temp_dir:
        yield Path(temp_dir)


class TestGenerateJSONTool:
    """Test JSON generation tool."""

    def test_tool_initialization(self, agent_with_data):
        """Test JSON tool can be initialized."""
        tool = GenerateJSONTool(agent_with_data.graph)
        assert tool.name == "generate_json"
        assert "JSON" in tool.description

    def test_generate_json_file(self, agent_with_data, temp_output_dir):
        """Test generating JSON export."""
        tool = GenerateJSONTool(agent_with_data.graph)
        output_path = temp_output_dir / "export.json"

        result = tool._run(str(output_path))

        assert output_path.exists()
        assert "successfully" in result.lower()

        # Verify JSON is valid
        with open(output_path) as f:
            data = json.load(f)

        assert "metadata" in data
        assert "nodes" in data
        assert "edges" in data

    def test_json_contains_graph_data(self, agent_with_data, temp_output_dir):
        """Test JSON export contains correct graph data."""
        tool = GenerateJSONTool(agent_with_data.graph)
        output_path = temp_output_dir / "export.json"

        tool._run(str(output_path))

        with open(output_path) as f:
            data = json.load(f)

        # Should have nodes and edges
        assert len(data["nodes"]) > 0
        assert len(data["edges"]) > 0

        # Metadata should have stats
        assert "total_nodes" in data["metadata"]
        assert "total_edges" in data["metadata"]


class TestGenerateCSVTool:
    """Test CSV generation tool."""

    def test_tool_initialization(self, agent_with_data):
        """Test CSV tool can be initialized."""
        tool = GenerateCSVTool(agent_with_data.graph)
        assert tool.name == "generate_csv"
        assert "CSV" in tool.description

    def test_generate_csv_file(self, agent_with_data, temp_output_dir):
        """Test generating CSV export."""
        tool = GenerateCSVTool(agent_with_data.graph)
        output_path = temp_output_dir / "export.csv"

        result = tool._run(str(output_path))

        assert output_path.exists()
        assert "successfully" in result.lower()

        # Verify CSV is valid
        df = pd.read_csv(output_path)
        assert len(df) > 0

    def test_csv_contains_lineage_data(self, agent_with_data, temp_output_dir):
        """Test CSV export contains lineage data."""
        tool = GenerateCSVTool(agent_with_data.graph)
        output_path = temp_output_dir / "lineage.csv"

        tool._run(str(output_path), export_type="lineage")

        df = pd.read_csv(output_path)

        # Should have source and target columns
        assert "source" in df.columns or "from_id" in df.columns
        assert "target" in df.columns or "to_id" in df.columns
        assert len(df) > 0

    def test_csv_export_nodes(self, agent_with_data, temp_output_dir):
        """Test CSV export of nodes."""
        tool = GenerateCSVTool(agent_with_data.graph)
        output_path = temp_output_dir / "nodes.csv"

        tool._run(str(output_path), export_type="nodes")

        df = pd.read_csv(output_path)

        # Should have node data
        assert "node_id" in df.columns or "id" in df.columns
        assert len(df) > 0


class TestGenerateExcelTool:
    """Test Excel generation tool."""

    def test_tool_initialization(self, agent_with_data):
        """Test Excel tool can be initialized."""
        tool = GenerateExcelTool(agent_with_data.graph)
        assert tool.name == "generate_excel"
        assert "Excel" in tool.description

    def test_generate_excel_file(self, agent_with_data, temp_output_dir):
        """Test generating Excel export."""
        tool = GenerateExcelTool(agent_with_data.graph)
        output_path = temp_output_dir / "report.xlsx"

        result = tool._run(str(output_path))

        assert output_path.exists()
        assert "successfully" in result.lower()

        # Verify Excel is valid
        wb = load_workbook(output_path)
        assert len(wb.sheetnames) > 0

    def test_excel_has_multiple_sheets(self, agent_with_data, temp_output_dir):
        """Test Excel export has multiple sheets."""
        tool = GenerateExcelTool(agent_with_data.graph)
        output_path = temp_output_dir / "report.xlsx"

        tool._run(str(output_path))

        wb = load_workbook(output_path)

        # Should have summary, nodes, edges sheets
        expected_sheets = ["Summary", "Nodes", "Edges"]
        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames

    def test_excel_summary_sheet_has_stats(self, agent_with_data, temp_output_dir):
        """Test Excel summary sheet contains statistics."""
        tool = GenerateExcelTool(agent_with_data.graph)
        output_path = temp_output_dir / "report.xlsx"

        tool._run(str(output_path))

        wb = load_workbook(output_path)
        ws = wb["Summary"]

        # Should have some data
        assert ws["A1"].value is not None


class TestGeneratePythonTool:
    """Test Python code generation tool."""

    def test_tool_initialization(self):
        """Test Python tool can be initialized."""
        tool = GeneratePythonTool()
        assert tool.name == "generate_python"
        assert "Python" in tool.description

    def test_generate_python_from_cobol(self, temp_output_dir):
        """Test generating Python from COBOL."""
        tool = GeneratePythonTool()

        # Simple COBOL snippet
        cobol_code = """
        IDENTIFICATION DIVISION.
        PROGRAM-ID. HELLO.
        PROCEDURE DIVISION.
            DISPLAY 'Hello World'.
            STOP RUN.
        """

        output_path = temp_output_dir / "hello.py"
        result = tool._run(cobol_code, str(output_path), source_type="cobol")

        assert output_path.exists()
        assert "successfully" in result.lower()

        # Verify Python is valid syntax
        with open(output_path) as f:
            code = f.read()
            compile(code, str(output_path), 'exec')  # Should not raise

    def test_generate_python_from_jcl(self, temp_output_dir):
        """Test generating Python from JCL."""
        tool = GeneratePythonTool()

        # Simple JCL snippet
        jcl_code = """
        //HELLO JOB
        //STEP1 EXEC PGM=IEFBR14
        """

        output_path = temp_output_dir / "job.py"
        result = tool._run(jcl_code, str(output_path), source_type="jcl")

        assert output_path.exists()
        assert "successfully" in result.lower()


class TestCodeGenerationIntegration:
    """Integration tests for all code generation tools."""

    def test_generate_all_formats(self, agent_with_data, temp_output_dir):
        """Test generating all export formats."""
        # JSON
        json_tool = GenerateJSONTool(agent_with_data.graph)
        json_path = temp_output_dir / "export.json"
        json_tool._run(str(json_path))
        assert json_path.exists()

        # CSV
        csv_tool = GenerateCSVTool(agent_with_data.graph)
        csv_path = temp_output_dir / "export.csv"
        csv_tool._run(str(csv_path))
        assert csv_path.exists()

        # Excel
        excel_tool = GenerateExcelTool(agent_with_data.graph)
        excel_path = temp_output_dir / "report.xlsx"
        excel_tool._run(str(excel_path))
        assert excel_path.exists()

        # All files should exist
        assert json_path.exists()
        assert csv_path.exists()
        assert excel_path.exists()

    def test_export_consistency(self, agent_with_data, temp_output_dir):
        """Test that JSON and CSV exports have consistent node counts."""
        # Generate JSON
        json_tool = GenerateJSONTool(agent_with_data.graph)
        json_path = temp_output_dir / "export.json"
        json_tool._run(str(json_path))

        # Generate CSV nodes
        csv_tool = GenerateCSVTool(agent_with_data.graph)
        csv_path = temp_output_dir / "nodes.csv"
        csv_tool._run(str(csv_path), export_type="nodes")

        # Load both
        with open(json_path) as f:
            json_data = json.load(f)

        csv_df = pd.read_csv(csv_path)

        # Node counts should match (approximately, accounting for format differences)
        assert abs(len(json_data["nodes"]) - len(csv_df)) <= 1
