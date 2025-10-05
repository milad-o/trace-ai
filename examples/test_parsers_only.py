"""Test all parsers without requiring API key - just graph operations."""

import json
from pathlib import Path

import openpyxl
import pandas as pd

from traceai.agents import EnterpriseAgent
from traceai.graph.queries import GraphQueries


def create_test_data():
    """Creates test data for all parsers."""
    test_dir = Path("test_data")
    test_dir.mkdir(exist_ok=True)

    # 1. JSON ETL config
    json_config = {
        "name": "SalesDataPipeline",
        "jobs": [
            {"id": "extract", "name": "Extract", "depends_on": []},
            {"id": "transform", "name": "Transform", "depends_on": ["extract"]},
        ],
    }
    json_file = test_dir / "pipeline.json"
    with open(json_file, "w") as f:
        json.dump(json_config, f)

    # 2. CSV lineage
    lineage_df = pd.DataFrame(
        {
            "source_table": ["raw_sales", "raw_customers"],
            "target_table": ["dim_sales", "dim_customers"],
            "transformation": ["Aggregate", "Deduplicate"],
        }
    )
    csv_file = test_dir / "lineage.csv"
    lineage_df.to_csv(csv_file, index=False)

    # 3. Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Product"
    ws["B1"] = "=SUM(A:A)"
    excel_file = test_dir / "report.xlsx"
    wb.save(excel_file)

    return test_dir


def test_all_parsers():
    """Test loading and parsing all file formats."""
    print("=" * 70)
    print("TESTING ALL PARSERS - No API Key Required")
    print("=" * 70)

    # Create test data
    test_dir = create_test_data()
    print(f"\n✓ Created test data in {test_dir}/")

    # Create agent without AI (graph-only mode)
    print("\n📚 Loading documents...")
    agent = EnterpriseAgent(model_provider="anthropic")  # Will work in graph-only mode

    # Load SSIS
    print("  Loading SSIS packages...")
    agent.load_documents("examples/sample_packages")

    # Load COBOL/JCL
    print("  Loading COBOL/JCL...")
    agent.load_documents("examples/sample_mainframe", pattern=["*.cbl", "*.jcl"])

    # Load JSON/CSV/Excel
    print("  Loading JSON/CSV/Excel...")
    agent.load_documents(str(test_dir), pattern=["*.json", "*.csv", "*.xlsx"])

    # Get statistics
    queries = GraphQueries(agent.graph)
    stats = queries.get_graph_stats()

    print("\n" + "=" * 70)
    print("GRAPH STATISTICS")
    print("=" * 70)
    print(f"Total Nodes: {stats['total_nodes']}")
    print(f"Total Edges: {stats['total_edges']}")
    print(f"\nNode Breakdown:")
    print(f"  Packages: {stats.get('package_nodes', 0)}")
    print(f"  Tasks: {stats.get('task_nodes', 0)}")
    print(f"  Tables: {stats.get('table_nodes', 0)}")
    print(f"  Connections: {stats.get('connection_nodes', 0)}")

    # Test each parser worked
    print("\n" + "=" * 70)
    print("PARSER VERIFICATION")
    print("=" * 70)

    parsers_tested = {
        "SSIS": False,
        "COBOL": False,
        "JCL": False,
        "JSON": False,
        "CSV": False,
        "Excel": False,
    }

    for node_id, node_data in agent.graph.nodes(data=True):
        node_type = node_data.get("type")
        name = node_data.get("name", "")

        # Check for each parser type
        if "CustomerETL" in name or "SalesAggregation" in name:
            parsers_tested["SSIS"] = True
        if ".cbl" in str(node_id) or "CUST" in name or "PAY" in name:
            parsers_tested["COBOL"] = True
        if ".jcl" in str(node_id) or "JCL" in name:
            parsers_tested["JCL"] = True
        if "SalesDataPipeline" in name or "Extract" in name:
            parsers_tested["JSON"] = True
        if "raw_sales" in str(node_id) or "dim_sales" in str(node_id):
            parsers_tested["CSV"] = True
        if "Excel" in str(node_data.get("description", "")):
            parsers_tested["Excel"] = True

    for parser, tested in parsers_tested.items():
        status = "✅" if tested else "❌"
        print(f"{status} {parser} Parser")

    # Test code generation tools
    print("\n" + "=" * 70)
    print("TESTING CODE GENERATION TOOLS")
    print("=" * 70)

    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)

    # Test JSON export
    print("\n1. Testing JSON Export...")
    from traceai.tools.code_generation_tools import GenerateJSONTool

    json_tool = GenerateJSONTool(queries=queries)
    result = json_tool._run(output_path="output/test_export.json", include_metadata=True)
    print(f"   {result}")

    # Test CSV export
    print("\n2. Testing CSV Export...")
    from traceai.tools.code_generation_tools import GenerateCSVTool

    csv_tool = GenerateCSVTool(queries=queries)
    result = csv_tool._run(output_path="output/test_lineage.csv", export_type="lineage")
    print(f"   {result}")

    # Test Excel export
    print("\n3. Testing Excel Export...")
    from traceai.tools.code_generation_tools import GenerateExcelTool

    excel_tool = GenerateExcelTool(queries=queries)
    result = excel_tool._run(
        output_path="output/test_analysis.xlsx", include_sheets=["summary", "nodes"]
    )
    print(f"   {result}")

    # Test Python generator
    print("\n4. Testing Python Code Generator...")
    from traceai.tools.python_generator import GeneratePythonTool

    python_tool = GeneratePythonTool(queries=queries)

    # Find a COBOL program
    cobol_node = None
    for node_id, node_data in agent.graph.nodes(data=True):
        if node_data.get("type") == "package" and "COBOL" in str(
            node_data.get("description", "")
        ):
            cobol_node = node_id
            break

    if cobol_node:
        result = python_tool._run(
            node_id=cobol_node, output_path="output/test_converted.py", style="script"
        )
        print(f"   {result[:200]}...")
    else:
        print("   ⚠️  No COBOL program found to convert")

    # Verify files created
    print("\n" + "=" * 70)
    print("GENERATED FILES")
    print("=" * 70)

    for file_path in [
        "output/test_export.json",
        "output/test_lineage.csv",
        "output/test_analysis.xlsx",
        "output/test_converted.py",
    ]:
        if Path(file_path).exists():
            size = Path(file_path).stat().st_size
            print(f"✅ {file_path} ({size:,} bytes)")
        else:
            print(f"❌ {file_path}")

    # Summary
    print("\n" + "=" * 70)
    print("TEST SUMMARY")
    print("=" * 70)

    parsers_ok = sum(parsers_tested.values())
    print(f"Parsers: {parsers_ok}/6 working")

    files_created = sum(
        1
        for f in ["output/test_export.json", "output/test_lineage.csv", "output/test_analysis.xlsx"]
        if Path(f).exists()
    )
    print(f"Code Gen Tools: {files_created}/3 working")

    if parsers_ok == 6 and files_created == 3:
        print("\n✅ ALL TESTS PASSED!")
    else:
        print(f"\n⚠️  Some tests failed")

    print("=" * 70)


if __name__ == "__main__":
    test_all_parsers()
